import webbrowser
import customtkinter as ctk
import smtplib
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import scrolledtext
from tkinter import ttk
from PIL import Image, ImageTk
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import os
import sys
import time
import concurrent.futures
from datetime import datetime
import threading
import queue

# Global variables
sender_email = ""
sender_password = ""
data_path = ""
template_path = ""
subject = ""
success_path = "done.txt"
fail_path = "fail.txt"
log_text = None  # Global variable for log text widget
progress_bar = None  # Global variable for progress bar
progress_label = None  # Global variable for progress label
log_queue = queue.Queue()  # Queue for thread-safe logging
is_sending = False  # Flag to track if sending is in progress


def login():
    global sender_email, sender_password
    # Get sender email and password
    sender_email = email_entry.get()
    sender_password = password_entry.get()

    if sender_email == "" or sender_password == "":
        messagebox.showerror("Lỗi", "Vui lòng nhập email và mật khẩu của bạn.")
        return
    # Try to login with email and password
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    try:
        server.login(sender_email, sender_password)
        screen1.pack_forget()
        screen2.pack(fill="both", expand=True)
    except Exception as e:
        # Delete email and password
        password_entry.delete(0, ctk.END)
        messagebox.showerror(
            "Lỗi", f"Đăng nhập thất bại. Vui lòng kiểm tra lại email và mật khẩu. {e}"
        )


def submitfile():
    global data_path, template_path
    data_path = data_path_entry.get()
    template_path = index_path_entry.get()
    if data_path == "" or template_path == "":
        messagebox.showerror(
            "Trường trống!", "Vui lòng nhập đường dẫn dữ liệu và đường dẫn mẫu."
        )
        return
    screen2.pack_forget()
    screen3.pack(fill="both", expand=True)


def log_message(message):
    """Add a message to the log queue for thread-safe logging"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    log_queue.put(f"[{timestamp}] {message}")


def process_log_queue():
    """Process log messages from queue and update UI"""
    global log_text
    try:
        while True:
            message = log_queue.get_nowait()
            if log_text:
                log_text.insert(tk.END, f"{message}\n")
                log_text.see(tk.END)
    except queue.Empty:
        pass
    finally:
        # Schedule next check
        app.after(100, process_log_queue)


def send_email():
    global subject, is_sending

    if is_sending:
        messagebox.showwarning("Đang gửi", "Đang gửi email, vui lòng chờ...")
        return

    subject = input_subject_entry.get()
    if subject == "":
        messagebox.showerror("Trường trống!", "Vui lòng nhập tiêu đề email.")
        return

    # Clear previous log
    log_text.delete(1.0, tk.END)

    # Reset progress bar
    if progress_bar:
        progress_bar["value"] = 0
        progress_label.config(text="Chuẩn bị gửi email...")

    # Disable send button
    send_button.configure(state="disabled", text="Đang gửi...")
    is_sending = True

    log_message("Bắt đầu gửi email...")

    # Run send in separate thread to avoid UI freezing
    thread = threading.Thread(target=send_email_thread, daemon=True)
    thread.start()


def send_email_thread():
    """Send emails in background thread"""
    global is_sending
    try:
        send()
        log_message("Hoàn thành gửi email!")

        # Schedule UI update in main thread
        app.after(
            0,
            lambda: messagebox.showinfo("Thành công", "Emails đã được gửi thành công!"),
        )
    except Exception as e:
        log_message(f"Lỗi: {str(e)}")
        app.after(0, lambda: messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}"))
    finally:
        # Always reset UI state even if an error occurs
        def reset_state():
            global is_sending
            send_button.configure(state="normal", text="Gửi Email")
            is_sending = False

        app.after(0, reset_state)


def back_to_screen1():
    screen2.pack_forget()
    screen1.pack(fill="both", expand=True)


def back_to_screen2():
    screen3.pack_forget()
    screen2.pack(fill="both", expand=True)


def user_manual():
    webbrowser.open(
        "https://docs.google.com/document/d/1I5GG2K_csg7bR1ZQf-5ldzEtqPN-6IJN7xn6MJDujWM/edit?usp=sharing"
    )


def send_email_task(row_idx, recipient_data, content_type, subject, email_content):
    # Find email from recipient_data - look for various possible email column names
    receiver_email = None
    for key in recipient_data.keys():
        if key.lower() in ["email", "e-mail", "mail"]:
            receiver_email = recipient_data[key]
            break

    if not receiver_email:
        # If still no email found, try to get the first value that looks like an email
        for value in recipient_data.values():
            if "@" in str(value):
                receiver_email = str(value)
                break

    if not receiver_email:
        log_message(f"✗ Không tìm thấy email trong dữ liệu: {recipient_data}")
        # Still return a 4-tuple so the caller logic remains consistent
        return (row_idx, "Unknown", "No Email", False)

    # Try to find name - look for various possible name columns
    recipient_name = receiver_email  # Default fallback
    for key in recipient_data.keys():
        key_lower = key.lower()
        if any(
            name_key in key_lower
            for name_key in ["name", "tên", "họ", "firstname", "lastname", "$name"]
        ):
            if recipient_data[key]:
                recipient_name = str(recipient_data[key])
                break
    try:
        # Connect to Gmail's SMTP server with timeout
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        server = smtplib.SMTP(smtp_server, smtp_port, timeout=30)
        server.starttls()
        server.login(sender_email, sender_password)

        # Create a MIMEMultipart object to create the email
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject

        # Replace all placeholders in email content
        personalized_content = email_content
        for placeholder, value in recipient_data.items():
            personalized_content = personalized_content.replace(placeholder, value)

        message.attach(MIMEText(personalized_content, content_type))

        # Send the email
        server.sendmail(sender_email, receiver_email, message.as_string())

        # Close the SMTP connection
        server.quit()

        log_message(f"✓ Gửi thành công: {recipient_name} ({receiver_email})")
        return (row_idx, recipient_name, receiver_email, True)
    except Exception as e:
        log_message(f"✗ Gửi thất bại: {recipient_name} ({receiver_email}) - {str(e)}")
        return (row_idx, recipient_name, receiver_email, False)


def send():
    # Determine the content type based on the file extension
    if template_path.endswith(".html"):
        content_type = "html"
    else:
        content_type = "plain"

    # Read the email content template
    log_message("Đang đọc template email...")
    with open(template_path, "r", encoding="utf-8") as template_file:
        email_content = template_file.read()

    # Read data from the Excel file
    log_message("Đang đọc dữ liệu từ file Excel...")
    wb = openpyxl.load_workbook(data_path)
    sheet = wb.active

    # Read header row to get placeholders
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    placeholders = [str(h).strip() if h else "" for h in header_row]

    # Find email column index by looking for email-related column names
    email_col_index = None
    for i, placeholder in enumerate(placeholders):
        placeholder_lower = placeholder.lower()
        if any(
            email_key in placeholder_lower for email_key in ["email", "e-mail", "mail"]
        ):
            email_col_index = i
            break

    # If still not found, look for columns that might contain @ symbol
    if email_col_index is None:
        for row in sheet.iter_rows(
            min_row=2, max_row=3, values_only=True
        ):  # Check first few rows
            for i, value in enumerate(row):
                if value and "@" in str(value):
                    email_col_index = i
                    log_message(
                        f"Tìm thấy cột email tại vị trí {i+1}: '{placeholders[i] if i < len(placeholders) else 'Unknown'}'"
                    )
                    break
            if email_col_index is not None:
                break

    if email_col_index is None:
        log_message("✗ Không tìm thấy cột email trong file Excel")
        log_message(f"Các cột có sẵn: {placeholders}")
        raise ValueError("Không tìm thấy cột email trong file Excel")

    # Get the list of recipients with all column data and row indices
    recipients = []  # list of tuples: (row_index, row_data_dict)
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=False), start=2
    ):
        email_value = (
            row[email_col_index].value if row[email_col_index] is not None else None
        )
        if email_value:  # Check if email exists
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(placeholders) and placeholders[i]:
                    row_data[placeholders[i]] = (
                        str(cell.value) if cell.value is not None else ""
                    )
            recipients.append((row_idx, row_data))

    log_message(f"Tìm thấy {len(recipients)} người nhận. Bắt đầu gửi email...")

    # Log some sample recipient data for debugging
    if recipients:
        try:
            sample_keys = list(recipients[0][1].keys())  # (row_idx, row_data_dict)
        except Exception:
            sample_keys = []
        log_message(f"Ví dụ dữ liệu người nhận đầu tiên: {sample_keys}")

    # Use ThreadPoolExecutor to send emails concurrently
    sent_successfully = []  # List of successfully sent emails
    failed_recipients = []  # List of failed emails
    total_recipients = len(recipients)
    completed_count = 0

    with concurrent.futures.ThreadPoolExecutor(
        max_workers=3  # Further reduced for stability
    ) as executor:
        futures = [
            executor.submit(
                send_email_task,
                row_idx,
                recipient_data,
                content_type,
                subject,
                email_content,
            )
            for (row_idx, recipient_data) in recipients
        ]
        for future in concurrent.futures.as_completed(futures):
            try:
                row_idx, recipient_name, receiver_email, success = future.result()
            except Exception as e:
                # If a task raised unexpectedly, log and continue; mark as failed if possible
                log_message(f"Task lỗi bất ngờ: {e}")
                continue
            completed_count += 1

            if success:
                sent_successfully.append((recipient_name, receiver_email))
                # Mark True in column E (5th column)
                sheet.cell(row=row_idx, column=5).value = True
            else:
                failed_recipients.append((recipient_name, receiver_email))
                # Mark False in column E (5th column)
                sheet.cell(row=row_idx, column=5).value = False

            # Update progress bar and log
            progress_percent = (completed_count / total_recipients) * 100
            if progress_bar:

                def update_progress():
                    progress_bar.configure(value=progress_percent)
                    progress_label.config(
                        text=f"Tiến độ: {completed_count}/{total_recipients} ({progress_percent:.1f}%)"
                    )

                app.after(0, update_progress)

            log_message(
                f"Tiến độ: {completed_count}/{total_recipients} ({progress_percent:.1f}%)"
            )

    # Save results to workbook
    try:
        wb.save(data_path)
        log_message("Đã lưu kết quả True/False vào cột E của file Excel")
    except Exception as e:
        log_message(f"Không thể lưu file Excel: {e}")

    # Log summary
    log_message(
        f"Kết quả: {len(sent_successfully)} thành công, {len(failed_recipients)} thất bại"
    )
    # No longer writing done.txt/fail.txt; results are in column E


app = ctk.CTk()
app.title("Công cụ Tự động hóa Email")
app.geometry("800x700")
app.resizable(True, True)

# Screen 1: input sender email and password
screen1 = ctk.CTkFrame(master=app)
screen1.pack(fill="both", expand=True)

# Set the background image (optional)
# script_directory = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
# image_path = os.path.join(script_directory, "assets/cover2.jpg")
# original_image = Image.open(image_path)
# resized_image = original_image.resize((800, 180))
# tk_image = ImageTk.PhotoImage(resized_image)
# canvas = ctk.CTkCanvas(master=screen1, width=800, height=200)
# canvas.create_image(0, 0, anchor=tk.NW, image=tk_image)
# canvas.pack()

text = ctk.CTkLabel(
    master=screen1,
    text="Công cụ gửi email tự động CLB Sổ Media",
    font=("Arial", 20, "bold"),
    text_color="#402E7A",
)
text.pack(pady=30)

email_label = ctk.CTkLabel(master=screen1, text="Email", font=("Arial", 14))
email_label.pack()

email_entry = ctk.CTkEntry(master=screen1, width=300, height=30, font=("Arial", 12))
email_entry.pack(pady=5)

password_label = ctk.CTkLabel(master=screen1, text="Mật khẩu", font=("Arial", 14))
password_label.pack()

password_entry = ctk.CTkEntry(
    master=screen1, show="*", width=300, height=30, font=("Arial", 12)
)
password_entry.pack(pady=5)

login_button = ctk.CTkButton(
    master=screen1,
    text="Đăng nhập",
    width=100,
    height=40,
    command=login,
    corner_radius=10,
    fg_color="#4B70F5",
)
login_button.pack(pady=20)

user_manual_button = ctk.CTkButton(
    master=screen1,
    text="Hướng dẫn sử dụng",
    width=100,
    height=40,
    command=user_manual,
    corner_radius=10,
    fg_color="#264653",
)
user_manual_button.pack()

# Screen 2: file selection
screen2 = ctk.CTkFrame(master=app)

back_button = ctk.CTkButton(
    master=screen2,
    text="Trở lại",
    width=70,
    height=30,
    command=back_to_screen1,
    corner_radius=10,
    fg_color="#264653",
)
back_button.pack(side=tk.TOP, padx=20, pady=20)

data_path_label = ctk.CTkLabel(
    screen2, text="Đường dẫn tới file xlsx", font=("Arial", 14)
)
data_path_label.pack()

data_path_entry = ctk.CTkEntry(screen2, width=300, height=30, font=("Arial", 12))
data_path_entry.pack(pady=10)

select_data_button = ctk.CTkButton(
    screen2,
    text="Chọn file",
    command=lambda: (
        data_path_entry.delete(0, ctk.END),
        data_path_entry.insert(0, filedialog.askopenfilename()),
    ),
    width=100,
    height=40,
    corner_radius=10,
    fg_color="#2a9d8f",
)
select_data_button.pack()

index_path_label = ctk.CTkLabel(
    screen2, text="Đường dẫn tới file nội dung (.html hoặc .txt)", font=("Arial", 14)
)
index_path_label.pack()

index_path_entry = ctk.CTkEntry(screen2, width=300, height=30, font=("Arial", 12))
index_path_entry.pack(pady=10)

select_index_button = ctk.CTkButton(
    screen2,
    text="Chọn file",
    command=lambda: (
        index_path_entry.delete(0, ctk.END),
        index_path_entry.insert(0, filedialog.askopenfilename()),
    ),
    width=100,
    height=40,
    corner_radius=10,
    fg_color="#2a9d8f",
)
select_index_button.pack()

next_button = ctk.CTkButton(
    master=screen2,
    text="Tiếp theo",
    width=100,
    height=40,
    command=submitfile,
    corner_radius=10,
    fg_color="#4B70F5",
)
next_button.pack(pady=20)

# Screen 3: input subject and send emails
screen3 = ctk.CTkFrame(master=app)

# Top frame for back button
top_frame = ctk.CTkFrame(master=screen3)
top_frame.pack(fill="x", padx=20, pady=10)

back_tosc2 = ctk.CTkButton(
    master=top_frame,
    text="Trở lại",
    width=70,
    height=30,
    command=back_to_screen2,
    corner_radius=10,
    fg_color="#264653",
)
back_tosc2.pack(side=tk.LEFT)

# Input frame
input_frame = ctk.CTkFrame(master=screen3)
input_frame.pack(fill="x", padx=20, pady=10)

input_subject_label = ctk.CTkLabel(
    input_frame, text="Nhập tiêu đề:", font=("Arial", 14)
)
input_subject_label.pack(pady=5)

input_subject_entry = ctk.CTkEntry(
    input_frame, width=400, height=30, font=("Arial", 12)
)
input_subject_entry.pack(pady=5)

send_button = ctk.CTkButton(
    input_frame,
    text="Gửi Email",
    width=120,
    height=40,
    command=send_email,
    corner_radius=10,
    fg_color="#4B70F5",
)
send_button.pack(pady=10)

# Log frame
log_frame = ctk.CTkFrame(master=screen3)
log_frame.pack(fill="both", expand=True, padx=20, pady=10)

log_label = ctk.CTkLabel(log_frame, text="Log gửi email:", font=("Arial", 14, "bold"))
log_label.pack(anchor="w", padx=10, pady=(10, 5))

# Progress frame
progress_frame = ctk.CTkFrame(master=log_frame)
progress_frame.pack(fill="x", padx=10, pady=5)

progress_label = tk.Label(
    progress_frame,
    text="Sẵn sàng gửi email",
    font=("Arial", 10),
    bg="#212121",
    fg="#ffffff",
)
progress_label.pack(side=tk.LEFT, padx=5)

progress_bar = ttk.Progressbar(
    progress_frame, mode="determinate", length=300, maximum=100
)
progress_bar.pack(side=tk.RIGHT, padx=5, pady=5)

# Create log text widget
log_text = scrolledtext.ScrolledText(
    log_frame,
    height=12,
    width=80,
    font=("Consolas", 10),
    bg="#2b2b2b",
    fg="#ffffff",
    insertbackground="#ffffff",
    selectbackground="#404040",
    wrap=tk.WORD,
)
log_text.pack(fill="both", expand=True, padx=10, pady=5)

# Bottom frame for buttons
bottom_frame = ctk.CTkFrame(master=screen3)
bottom_frame.pack(fill="x", padx=20, pady=10)

done_button = ctk.CTkButton(
    master=bottom_frame,
    text="Hoàn thành",
    width=100,
    height=40,
    command=app.quit,
    corner_radius=10,
    fg_color="#264653",
)
done_button.pack(side=tk.RIGHT, padx=10)

to_success = ctk.CTkLabel(
    master=bottom_frame,
    text="Kết quả gửi email được lưu trong cột E của file Excel",
    font=("Arial", 12),
    text_color="#2a9d8f",
)
to_success.pack(side=tk.LEFT, padx=10)

# Start log processing
process_log_queue()

app.mainloop()
