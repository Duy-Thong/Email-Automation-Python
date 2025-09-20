import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import time

data_path = "data.xlsx"
template_path = "template.html"
sender_email = ""
sender_password = ""
subject = ""


def send_email(data_path, template_path, sender_email, sender_password, subject):
    # Determine the content type based on the file extension
    if template_path.endswith(".html"):
        content_type = "html"
    else:
        content_type = "plain"

    # Read the email content template
    with open(template_path, "r", encoding="utf-8") as template_file:
        email_content = template_file.read()

    # Read data from the Excel file
    wb = openpyxl.load_workbook(data_path)
    sheet = wb.active

    # Read header row to get placeholders
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    placeholders = [str(h).strip() if h else "" for h in header_row]

    # Find email column index by looking for "Email" placeholder
    email_col_index = None
    for i, placeholder in enumerate(placeholders):
        if placeholder.lower() == "email":
            email_col_index = i
            break

    if email_col_index is None:
        raise ValueError("Không tìm thấy cột 'Email' trong file Excel")

    # Get the list of recipients with all column data
    recipients = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[email_col_index]:  # Check if email exists
            row_data = {}
            for i, value in enumerate(row):
                if i < len(placeholders) and placeholders[i]:
                    row_data[placeholders[i]] = str(value) if value else ""
            recipients.append(row_data)

    # Connect to Gmail's SMTP server
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    try:
        # Log in to the email account
        server.login(sender_email, sender_password)
        # Send email to each recipient
        sent_successfully = []  # List of successfully sent emails
        failed_recipients = []  # List of failed emails

        for recipient_data in recipients:
            receiver_email = recipient_data["Email"]
            try:
                # Create a MIMEMultipart object to create the email
                message = MIMEMultipart()
                message["From"] = sender_email
                message["To"] = receiver_email
                message["Subject"] = subject

                # Replace all placeholders in email content
                personalized_content = email_content
                for placeholder, value in recipient_data.items():
                    personalized_content = personalized_content.replace(
                        placeholder, value
                    )

                # Attach the email content based on the determined content type
                message.attach(MIMEText(personalized_content, content_type))

                # Send the email
                server.sendmail(sender_email, receiver_email, message.as_string())

                # Record the successfully sent email
                recipient_name = recipient_data.get(
                    "$NAME", receiver_email
                )  # Use $NAME or email as fallback
                sent_successfully.append((recipient_name, receiver_email))
                print(f"{recipient_name}, {receiver_email}")
            except Exception as e:
                # Record the failed email
                recipient_name = recipient_data.get("$NAME", receiver_email)
                failed_recipients.append((recipient_name, receiver_email))
            time.sleep(1)

        # Close the SMTP connection
        server.quit()
    except Exception as e:
        print(f"Failed to send emails due to: {e}")
        # Close the SMTP connection in case of failure
        server.quit()
